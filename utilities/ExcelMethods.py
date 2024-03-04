import openpyxl


def numRows(file_name, sheet_name):
    Excel_File = openpyxl.load_workbook(file_name)  # loading the file
    Sheet = Excel_File[sheet_name]  # loading sheet name
    return Sheet.max_row  # In that specific we are getting max_rows


def readData(file_name, sheet_name, row_num, col_num):
    Excel_File = openpyxl.load_workbook(file_name)  # loading the file
    Sheet = Excel_File[sheet_name]  # loading sheet name
    return Sheet.cell(row=row_num, column=col_num).value  # returning the value of given row and column


def writeData(file_name, sheet_name, row_num, col_num, data):
    Excel_File = openpyxl.load_workbook(file_name)  # loading the file
    Sheet = Excel_File[sheet_name]  # loading sheet name
    Sheet.cell(row=row_num, column=col_num).value = data  # giving the value for given row and column
    Excel_File.save(file_name) # save the file

# readData(Test_data, LoginData,2,3)
